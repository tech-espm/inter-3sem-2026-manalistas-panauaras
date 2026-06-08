# -*- coding: utf-8 -*-
import os
import warnings
import mysql.connector
from dotenv import load_dotenv

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()
warnings.filterwarnings("ignore")

OUTPUT_PATH = os.path.join("public", "relatorios", "dados_brutos_com_zonas.xlsx")

STYLE_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FILL_HEADER = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FILL_HEADER2 = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
FILL_ZONA_1 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
FILL_ZONA_2 = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
FILL_ZONA_3 = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
FILL_ZONA_4 = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
FILL_ZONA_5 = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")
FILL_ZONA_MAP = {
    "UTI Adulto - Ala A": FILL_ZONA_1,
    "Sala de Espera - Triagem": FILL_ZONA_2,
    "Enfermaria - Ala B": FILL_ZONA_3,
    "Centro Cirurgico 02": FILL_ZONA_4,
    "Farmacia Central": FILL_ZONA_5,
}
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def conectar():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST", "localhost"),
        user=os.getenv("DB_USER", "root"),
        password=os.getenv("DB_PASSWORD", "root"),
        database=os.getenv("DB_DATABASE", "sensores_db"),
        port=int(os.getenv("DB_PORT", 3306)),
    )


def escrever_sheet(ws, df, sheet_title, col_width=18):
    ws.title = sheet_title
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = STYLE_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        # Aplica cor por zona se a coluna existir
        if "nome_setor" in df.columns:
            zona = row.get("nome_setor")
            fill = FILL_ZONA_MAP.get(zona)
            if fill:
                for c_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill

    for c_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col_name)),
        )
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(
            max_len + 3, 50
        )


def exportar():
    db = conectar()

    print("Carregando dados do banco...")

    creative = pd.read_sql(
        """
        SELECT
            c.id, c.data, c.id_sensor, sn.codigo_sensor, sn.tipo_sensor,
            st.id_setor, st.nome_setor, st.tipo_setor AS tipo_setor_nome,
            st.andar, st.capacidade_maxima,
            c.delta, c.luminosidade, c.umidade, c.temperatura,
            c.voc, c.co2, c.pressao_ar, c.ruido,
            c.aerosol_parado, c.aerosol_risco, c.ponto_orvalho
        FROM creative c
        JOIN sensores sn ON c.id_sensor = sn.id_sensor
        JOIN setores st ON sn.id_setor = st.id_setor
        ORDER BY c.data
        """,
        db,
    )
    print(f"  Creative: {len(creative)} registros")

    pca = pd.read_sql(
        """
        SELECT
            p.id, p.data, p.id_sensor, sn.codigo_sensor, sn.tipo_sensor,
            st.id_setor, st.nome_setor, st.tipo_setor AS tipo_setor_nome,
            st.andar, st.capacidade_maxima,
            p.delta, p.pessoas, p.luminosidade, p.umidade, p.temperatura
        FROM pca p
        JOIN sensores sn ON p.id_sensor = sn.id_sensor
        JOIN setores st ON sn.id_setor = st.id_setor
        ORDER BY p.data
        """,
        db,
    )
    print(f"  PCA: {len(pca)} registros")

    alertas = pd.read_sql(
        """
        SELECT
            a.id_alerta, a.id_sensor, a.id_setor, st.nome_setor, st.tipo_setor,
            a.parametro, a.valor_medido, a.limite_referencia, a.unidade,
            a.severidade, a.status, a.descricao,
            a.disparado_em, a.atendimento_iniciado_em, a.normalizado_em
        FROM alertas a
        JOIN setores st ON a.id_setor = st.id_setor
        ORDER BY a.disparado_em DESC
        """,
        db,
    )
    print(f"  Alertas: {len(alertas)} registros")

    limiares = pd.read_sql("SELECT * FROM limiares_ambiente ORDER BY tipo_setor, parametro", db)
    print(f"  Limiares: {len(limiares)} registros")

    setores = pd.read_sql("SELECT * FROM setores ORDER BY id_setor", db)
    print(f"  Setores: {len(setores)} registros")

    sensores = pd.read_sql(
        """
        SELECT sn.*, st.nome_setor, st.tipo_setor, st.andar
        FROM sensores sn
        JOIN setores st ON sn.id_setor = st.id_setor
        ORDER BY sn.id_sensor
        """,
        db,
    )
    print(f"  Sensores: {len(sensores)} registros")

    db.close()

    wb = Workbook()

    print("\nGerando Excel...")
    ws = wb.active
    escrever_sheet(ws, creative, "Dados Creative")
    print("  Sheet 1/6: Dados Creative")

    ws2 = wb.create_sheet()
    escrever_sheet(ws2, pca, "Dados PCA")
    print("  Sheet 2/6: Dados PCA")

    ws3 = wb.create_sheet()
    escrever_sheet(ws3, alertas, "Alertas")
    print("  Sheet 3/6: Alertas")

    ws4 = wb.create_sheet()
    escrever_sheet(ws4, limiares, "Limiares")
    print("  Sheet 4/6: Limiares")

    ws5 = wb.create_sheet()
    escrever_sheet(ws5, setores, "Setores")
    print("  Sheet 5/6: Setores")

    ws6 = wb.create_sheet()
    escrever_sheet(ws6, sensores, "Sensores")
    print("  Sheet 6/6: Sensores")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nArquivo salvo: {os.path.abspath(OUTPUT_PATH)}")
    print("Cada registro inclui as colunas de zona (nome_setor, tipo_setor, andar)")


if __name__ == "__main__":
    exportar()
