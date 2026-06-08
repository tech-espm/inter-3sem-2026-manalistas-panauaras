# -*- coding: utf-8 -*-
import os
import warnings
import mysql.connector
from dotenv import load_dotenv
from datetime import datetime, timedelta

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.label import DataLabelList

load_dotenv()
warnings.filterwarnings("ignore")

OUTPUT_PATH = os.path.join("public", "relatorios", "eda_completa.xlsx")

THRESHOLDS = {
    "CO2": {"critico": 1000, "atencao": 800, "unidade": "ppm"},
    "VOC": {"critico": 400, "atencao": 250, "unidade": "ppb"},
    "TEMPERATURA": {"critico": 26, "atencao": 24, "unidade": "C"},
    "UMIDADE": {"critico": 70, "atencao": 60, "unidade": "%"},
    "RUIDO": {"critico": 70, "atencao": 55, "unidade": "dB"},
}

THRESHOLDS_UTI = {
    "CO2": {"critico": 800, "atencao": 600, "unidade": "ppm"},
    "TEMPERATURA": {"critico": 24, "atencao": 22, "unidade": "C"},
    "UMIDADE": {"critico": 60, "atencao": 50, "unidade": "%"},
    "RUIDO": {"critico": 55, "atencao": 45, "unidade": "dB"},
}

STYLE_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
STYLE_TITLE = Font(name="Calibri", bold=True, size=14, color="0F172A")
STYLE_SUBTITLE = Font(name="Calibri", bold=True, size=12, color="1E40AF")
FILL_HEADER = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FILL_HEADER2 = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
FILL_LIGHT = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_GREEN = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
FILL_RED = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def conectar():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST", "localhost"),
        user=os.getenv("DB_USER", "root"),
        password=os.getenv("DB_PASSWORD", "root"),
        database=os.getenv("DB_DATABASE", "sensores_db"),
        port=int(os.getenv("DB_PORT", 3306)),
    )


def fmt_header_row(ws, row, max_col, fill=None):
    if fill is None:
        fill = FILL_HEADER
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = STYLE_HEADER
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def fmt_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if r % 2 == 0:
                cell.fill = FILL_LIGHT


def auto_width(ws, max_col, max_width=40, start_row=1):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=start_row, column=col)
        try:
            letter = cell.column_letter
        except AttributeError:
            continue
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, max_width)


def add_title(ws, title, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = STYLE_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def count_data_rows(ws, start_row, max_col):
    count = 0
    for r in range(start_row, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            count += 1
        else:
            break
    return count


def add_subtitle(ws, title, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = STYLE_SUBTITLE
    return row + 1


def desc_stats(series):
    s = series.dropna()
    if len(s) == 0:
        return {}
    return {
        "count": len(s),
        "media": round(s.mean(), 2),
        "mediana": round(s.median(), 2),
        "desvio_padrao": round(s.std(), 2),
        "min": round(s.min(), 2),
        "max": round(s.max(), 2),
        "p25": round(s.quantile(0.25), 2),
        "p75": round(s.quantile(0.75), 2),
        "amplitude": round(s.max() - s.min(), 2),
        "cv": round(s.std() / s.mean() * 100, 2) if s.mean() != 0 else 0,
    }


def gerar_eda():
    db = conectar()
    cursor = db.cursor(dictionary=True)

    print("[1/8] Carregando dados do banco...")
    creative = pd.read_sql("""
        SELECT c.*, s.id_setor, st.nome_setor, st.tipo_setor, st.andar
        FROM creative c
        JOIN sensores s ON c.id_sensor = s.id_sensor
        JOIN setores st ON s.id_setor = st.id_setor
        ORDER BY c.data
    """, db)
    creative["data"] = pd.to_datetime(creative["data"])
    creative["hora"] = creative["data"].dt.hour
    creative["dia_semana"] = creative["data"].dt.dayofweek
    creative["data_dia"] = creative["data"].dt.date

    pca = pd.read_sql("""
        SELECT p.*, s.id_setor, st.nome_setor, st.tipo_setor, st.andar, st.capacidade_maxima
        FROM pca p
        JOIN sensores s ON p.id_sensor = s.id_sensor
        JOIN setores st ON s.id_setor = st.id_setor
        ORDER BY p.data
    """, db)
    pca["data"] = pd.to_datetime(pca["data"])
    pca["hora"] = pca["data"].dt.hour
    pca["dia_semana"] = pca["data"].dt.dayofweek
    pca["data_dia"] = pca["data"].dt.date
    pca["pct_ocupacao"] = (pca["pessoas"] / pca["capacidade_maxima"] * 100).round(1)

    alertas = pd.read_sql("""
        SELECT a.*, st.nome_setor, st.tipo_setor
        FROM alertas a
        JOIN setores st ON a.id_setor = st.id_setor
        ORDER BY a.disparado_em DESC
    """, db)
    alertas["disparado_em"] = pd.to_datetime(alertas["disparado_em"])
    alertas["atendimento_iniciado_em"] = pd.to_datetime(alertas["atendimento_iniciado_em"])
    alertas["normalizado_em"] = pd.to_datetime(alertas["normalizado_em"])
    if not alertas.empty:
        alertas["tempo_resposta_min"] = (
            alertas["atendimento_iniciado_em"] - alertas["disparado_em"]
        ).dt.total_seconds() / 60
        alertas["tempo_resolucao_min"] = (
            alertas["normalizado_em"] - alertas["disparado_em"]
        ).dt.total_seconds() / 60

    setores = pd.read_sql("SELECT * FROM setores", db)
    sensores_tbl = pd.read_sql("SELECT * FROM sensores", db)
    limiares = pd.read_sql("SELECT * FROM limiares_ambiente", db)

    db.close()
    print("  OK - Creative: {} registros | PCA: {} | Alertas: {}".format(len(creative), len(pca), len(alertas)))

    wb = Workbook()

    # ==================================================================
    # SHEET 1 - VISÃO GERAL (DASHBOARD)
    # ==================================================================
    print("[2/8] Criando sheet: Visao Geral...")
    ws = wb.active
    ws.title = "Visao Geral"
    r = add_title(ws, "DASHBOARD - ANALISE EXPLORATORIA (EDA)")
    r += 1

    # KPIs
    kpis = [
        ("Total Registros Creative", len(creative)),
        ("Total Registros PCA", len(pca)),
        ("Sensores Ativos", len(sensores_tbl[sensores_tbl["ativo"] == 1])),
        ("Setores Monitorados", len(setores)),
        ("Total Alertas", len(alertas)),
        ("Alertas Abertos", len(alertas[alertas["status"] == "ABERTO"])),
        ("Alertas Criticos", len(alertas[alertas["severidade"] == "CRITICO"])),
        ("Periodo (Creative)", "{} a {}".format(
            creative["data"].min().strftime("%d/%m/%Y %H:%M") if not creative.empty else "-",
            creative["data"].max().strftime("%d/%m/%Y %H:%M") if not creative.empty else "-"
        )),
    ]
    ws.cell(row=r, column=1, value="INDICADOR").font = STYLE_HEADER
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.cell(row=r, column=2, value="VALOR").font = STYLE_HEADER
    ws.cell(row=r, column=2).fill = FILL_HEADER
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2).border = THIN_BORDER
    r += 1
    for label, val in kpis:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=val).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        r += 1

    r += 2
    # Medias globais creative
    r = add_subtitle(ws, "Medias Globais - Multisensor (Creative)", r)
    r += 1
    medias_creative = creative[["temperatura", "umidade", "co2", "voc", "ruido", "luminosidade", "pressao_ar", "ponto_orvalho"]].mean().round(2)
    ws.cell(row=r, column=1, value="Parametro").font = STYLE_HEADER
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.cell(row=r, column=2, value="Media").font = STYLE_HEADER
    ws.cell(row=r, column=2).fill = FILL_HEADER
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2).border = THIN_BORDER
    r += 1
    for param, val in medias_creative.items():
        ws.cell(row=r, column=1, value=param.capitalize()).border = THIN_BORDER
        ws.cell(row=r, column=2, value=val).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        r += 1

    auto_width(ws, 2, start_row=3)

    # ==================================================================
    # SHEET 2 - ESTATISTICAS DESCRITIVAS CREATIVE
    # ==================================================================
    print("[3/8] Criando sheet: Estatisticas Creative...")
    ws = wb.create_sheet("Estatisticas Creative")
    r = add_title(ws, "ESTATISTICAS DESCRITIVAS - MULTISENSOR (CREATIVE)")
    r += 1

    num_cols = ["temperatura", "umidade", "co2", "voc", "ruido", "luminosidade", "pressao_ar", "ponto_orvalho", "delta", "aerosol_parado", "aerosol_risco"]
    stats_data = {}
    for col in num_cols:
        if col in creative.columns:
            stats_data[col] = desc_stats(creative[col])

    df_stats = pd.DataFrame(stats_data).T
    df_stats.index.name = "parametro"
    df_stats = df_stats.reset_index()

    for c_idx, col_name in enumerate(df_stats.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
    fmt_header_row(ws, r, len(df_stats.columns))
    r += 1

    for _, row_data in df_stats.iterrows():
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        r += 1
    fmt_data_area(ws, r - len(df_stats), r - 1, len(df_stats.columns))
    auto_width(ws, len(df_stats.columns))

    # ==================================================================
    # SHEET 3 - ESTATISTICAS DESCRITIVAS PCA
    # ==================================================================
    print("[4/8] Criando sheet: Estatisticas PCA...")
    ws = wb.create_sheet("Estatisticas PCA")
    r = add_title(ws, "ESTATISTICAS DESCRITIVAS - OCUPACAO (PCA/HPD2)")
    r += 1

    num_cols_pca = ["pessoas", "temperatura", "umidade", "luminosidade", "delta", "pct_ocupacao"]
    stats_data_pca = {}
    for col in num_cols_pca:
        if col in pca.columns:
            stats_data_pca[col] = desc_stats(pca[col])

    df_stats_pca = pd.DataFrame(stats_data_pca).T
    df_stats_pca.index.name = "parametro"
    df_stats_pca = df_stats_pca.reset_index()

    for c_idx, col_name in enumerate(df_stats_pca.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
    fmt_header_row(ws, r, len(df_stats_pca.columns))
    r += 1
    for _, row_data in df_stats_pca.iterrows():
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        r += 1
    fmt_data_area(ws, r - len(df_stats_pca), r - 1, len(df_stats_pca.columns))
    auto_width(ws, len(df_stats_pca.columns))

    # ==================================================================
    # SHEET 4 - SERIE TEMPORAL (agregada por hora)
    # ==================================================================
    print("[5/8] Criando sheet: Serie Temporal...")
    ws = wb.create_sheet("Serie Temporal")

    creative_hora = creative.set_index("data").resample("1H")[
        ["temperatura", "umidade", "co2", "voc", "ruido"]
    ].mean().round(2).dropna().reset_index()
    creative_hora["data"] = creative_hora["data"].dt.strftime("%Y-%m-%d %H:%M")

    r = add_title(ws, "SERIE TEMPORAL (MEDIA POR HORA) - MULTISENSOR")
    r += 1
    for c_idx, col_name in enumerate(creative_hora.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
    fmt_header_row(ws, r, len(creative_hora.columns))
    r += 1

    for _, row_data in creative_hora.iterrows():
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        r += 1
    fmt_data_area(ws, r - len(creative_hora), r - 1, len(creative_hora.columns))

    # Grafico de linhas
    if len(creative_hora) > 1:
        chart = LineChart()
        chart.title = "Tendencia CO2 e Temperatura (Media Horaria)"
        chart.style = 10
        chart.y_axis.title = "Valor"
        chart.x_axis.title = "Data/Hora"
        chart.height = 15
        chart.width = 30

        data_start = r - len(creative_hora)
        data_end = r - 1

        co2_ref = Reference(ws, min_col=4, min_row=data_start, max_col=4, max_row=data_end)
        temp_ref = Reference(ws, min_col=2, min_row=data_start, max_col=2, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end)

        chart.add_data(co2_ref, titles_from_data=True)
        chart.add_data(temp_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.width = 20000
        chart.series[1].graphicalProperties.line.width = 20000
        ws.add_chart(chart, "K2")

    auto_width(ws, len(creative_hora.columns))

    # ==================================================================
    # SHEET 5 - ANALISE POR SETOR
    # ==================================================================
    print("[6/8] Criando sheet: Analise por Setor...")
    ws = wb.create_sheet("Analise por Setor")

    r = add_title(ws, "ANALISE AGREGADA POR SETOR")
    r += 1

    # Creative por setor
    r = add_subtitle(ws, "Multisensor por Setor", r)
    r += 1
    setor_creative = creative.groupby("nome_setor").agg({
        "temperatura": ["mean", "std", "min", "max"],
        "umidade": ["mean", "std", "min", "max"],
        "co2": ["mean", "std", "min", "max"],
        "voc": ["mean", "std", "min", "max"],
        "ruido": ["mean", "std", "min", "max"],
    }).round(2)
    setor_creative.columns = ["_".join(c).strip() for c in setor_creative.columns]
    setor_creative = setor_creative.reset_index()

    for c_idx, col_name in enumerate(setor_creative.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
    fmt_header_row(ws, r, len(setor_creative.columns))
    r += 1
    for _, row_data in setor_creative.iterrows():
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        r += 1
    fmt_data_area(ws, r - len(setor_creative), r - 1, len(setor_creative.columns))

    r += 2
    r = add_subtitle(ws, "Ocupacao (PCA) por Setor", r)
    r += 1
    setor_pca = pca.groupby("nome_setor").agg({
        "pessoas": ["mean", "std", "min", "max", "sum"],
        "pct_ocupacao": ["mean", "max"],
        "temperatura": ["mean", "std"],
        "umidade": ["mean", "std"],
    }).round(2)
    setor_pca.columns = ["_".join(c).strip() for c in setor_pca.columns]
    setor_pca = setor_pca.reset_index()

    for c_idx, col_name in enumerate(setor_pca.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
    fmt_header_row(ws, r, len(setor_pca.columns))
    r += 1
    for _, row_data in setor_pca.iterrows():
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        r += 1
    fmt_data_area(ws, r - len(setor_pca), r - 1, len(setor_pca.columns))

    auto_width(ws, max(len(setor_creative.columns), len(setor_pca.columns)))

    # ==================================================================
    # SHEET 6 - MATRIZ DE CORRELACAO
    # ==================================================================
    print("[7/8] Criando sheet: Correlacao...")
    ws = wb.create_sheet("Correlacao")

    r = add_title(ws, "MATRIZ DE CORRELACAO - VARIAVEIS AMBIENTAIS")
    r += 1

    corr_cols = ["temperatura", "umidade", "co2", "voc", "ruido", "luminosidade", "pressao_ar", "ponto_orvalho"]
    existing = [c for c in corr_cols if c in creative.columns]
    corr_matrix = creative[existing].corr().round(4)

    vars_list = list(corr_matrix.columns)
    ws.cell(row=r, column=1, value="Parametro")
    ws.cell(row=r, column=1).font = STYLE_HEADER
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.cell(row=r, column=1).border = THIN_BORDER
    for j, var in enumerate(vars_list, 2):
        ws.cell(row=r, column=j, value=var)
        ws.cell(row=r, column=j).font = STYLE_HEADER
        ws.cell(row=r, column=j).fill = FILL_HEADER
        ws.cell(row=r, column=j).border = THIN_BORDER
    r += 1

    for i, var_i in enumerate(vars_list):
        ws.cell(row=r, column=1, value=var_i)
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for j, var_j in enumerate(vars_list, 2):
            val = corr_matrix.loc[var_i, var_j]
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            cell.number_format = "0.0000"
            if abs(val) > 0.7:
                cell.fill = FILL_GREEN if val > 0 else FILL_RED
            elif abs(val) > 0.4:
                cell.fill = FILL_YELLOW
        r += 1

    r += 2
    r = add_subtitle(ws, "Interpretacao (|r| > 0.7 = verde/vermelho, |r| > 0.4 = amarelo)", r)

    r += 2
    r = add_subtitle(ws, "Correlacao Pessoas vs Variaveis Ambientais", r)
    r += 1

    if not creative.empty and not pca.empty:
        creative_agg = creative.groupby("id_sensor").agg({
            "co2": "mean", "temperatura": "mean", "umidade": "mean", "ruido": "mean"
        }).round(2).reset_index()
        pca_agg = pca.groupby("id_sensor").agg({
            "pessoas": "mean", "pct_ocupacao": "mean"
        }).round(2).reset_index()
        merged = pd.merge(creative_agg, pca_agg, on="id_sensor", how="inner")
        if len(merged) > 1:
            cross_corr = merged[["co2", "temperatura", "umidade", "ruido", "pessoas"]].corr().round(4)
            cross_vars = list(cross_corr.columns)
            ws.cell(row=r, column=1, value="Parametro")
            ws.cell(row=r, column=1).font = STYLE_HEADER
            ws.cell(row=r, column=1).fill = FILL_HEADER2
            ws.cell(row=r, column=1).border = THIN_BORDER
            for j, var in enumerate(cross_vars, 2):
                ws.cell(row=r, column=j, value=var)
                ws.cell(row=r, column=j).font = STYLE_HEADER
                ws.cell(row=r, column=j).fill = FILL_HEADER2
                ws.cell(row=r, column=j).border = THIN_BORDER
            r += 1
            for i, var_i in enumerate(cross_vars):
                ws.cell(row=r, column=1, value=var_i)
                ws.cell(row=r, column=1).border = THIN_BORDER
                for j, var_j in enumerate(cross_vars, 2):
                    val = cross_corr.loc[var_i, var_j]
                    cell = ws.cell(row=r, column=j, value=val)
                    cell.border = THIN_BORDER
                    cell.number_format = "0.0000"
                r += 1

    auto_width(ws, len(vars_list) + 1)

    # ==================================================================
    # SHEET 7 - ANALISE DE ALERTAS
    # ==================================================================
    print("[8/8] Criando sheet: Alertas...")
    ws = wb.create_sheet("Alertas")

    r = add_title(ws, "ANALISE DE ALERTAS E SLA")
    r += 1

    # KPIs de alertas
    total_alertas = len(alertas)
    if total_alertas > 0:
        kpi_data = [
            ("Total Alertas", total_alertas),
            ("Criticos", len(alertas[alertas["severidade"] == "CRITICO"])),
            ("Atencao", len(alertas[alertas["severidade"] == "ATENCAO"])),
            ("Abertos", len(alertas[alertas["status"] == "ABERTO"])),
            ("Em Atendimento", len(alertas[alertas["status"] == "EM_ATENDIMENTO"])),
            ("Resolvidos", len(alertas[alertas["status"] == "RESOLVIDO"])),
            ("Tempo Medio Resposta (min)", round(alertas["tempo_resposta_min"].dropna().mean(), 1)),
            ("Tempo Medio Resolucao (min)", round(alertas["tempo_resolucao_min"].dropna().mean(), 1)),
        ]
        ws.cell(row=r, column=1, value="KPI").font = STYLE_HEADER
        ws.cell(row=r, column=1).fill = FILL_HEADER
        ws.cell(row=r, column=2, value="Valor").font = STYLE_HEADER
        ws.cell(row=r, column=2).fill = FILL_HEADER
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER
        r += 1
        for label, val in kpi_data:
            ws.cell(row=r, column=1, value=label).border = THIN_BORDER
            ws.cell(row=r, column=2, value=val).border = THIN_BORDER
            ws.cell(row=r, column=1).alignment = ALIGN_LEFT
            ws.cell(row=r, column=2).alignment = ALIGN_CENTER
            r += 1

        r += 1
        r = add_subtitle(ws, "Alertas por Parametro", r)
        r += 1
        param_counts = alertas["parametro"].value_counts().reset_index()
        param_counts.columns = ["Parametro", "Quantidade"]
        for c_idx, col_name in enumerate(param_counts.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name)
        fmt_header_row(ws, r, 2)
        r += 1
        for _, row_data in param_counts.iterrows():
            ws.cell(row=r, column=1, value=row_data["Parametro"]).border = THIN_BORDER
            ws.cell(row=r, column=2, value=row_data["Quantidade"]).border = THIN_BORDER
            ws.cell(row=r, column=2).alignment = ALIGN_CENTER
            r += 1

        r += 1
        r = add_subtitle(ws, "Alertas por Setor", r)
        r += 1
        setor_counts = alertas["nome_setor"].value_counts().reset_index()
        setor_counts.columns = ["Setor", "Quantidade"]
        for c_idx, col_name in enumerate(setor_counts.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name)
        fmt_header_row(ws, r, 2)
        r += 1
        for _, row_data in setor_counts.iterrows():
            ws.cell(row=r, column=1, value=row_data["Setor"]).border = THIN_BORDER
            ws.cell(row=r, column=2, value=row_data["Quantidade"]).border = THIN_BORDER
            ws.cell(row=r, column=2).alignment = ALIGN_CENTER
            r += 1

        r += 1
        r = add_subtitle(ws, "Log Completo de Alertas", r)
        r += 1
        log_cols = ["disparado_em", "nome_setor", "parametro", "valor_medido", "severidade", "status", "tempo_resposta_min", "tempo_resolucao_min"]
        # Filtrar colunas existentes
        log_cols = [c for c in log_cols if c in alertas.columns]
        alertas_log = alertas[log_cols].copy()
        if "disparado_em" in alertas_log.columns:
            alertas_log["disparado_em"] = alertas_log["disparado_em"].dt.strftime("%Y-%m-%d %H:%M")

        for c_idx, col_name in enumerate(alertas_log.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
        fmt_header_row(ws, r, len(alertas_log.columns))
        r += 1
        for _, row_data in alertas_log.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c_idx, value=val)
            r += 1
        fmt_data_area(ws, r - len(alertas_log), r - 1, len(alertas_log.columns))

    auto_width(ws, max(8, len(alertas_log.columns) if total_alertas > 0 else 2))

    # ==================================================================
    # SHEET 8 - PADROES HORARIOS
    # ==================================================================
    ws = wb.create_sheet("Padroes Horarios")

    r = add_title(ws, "PADROES POR HORA DO DIA")
    r += 1

    if not creative.empty:
        r = add_subtitle(ws, "Media Horaria - Multisensor", r)
        r += 1
        hora_creative = creative.groupby("hora")[
            ["temperatura", "umidade", "co2", "voc", "ruido"]
        ].mean().round(2).reset_index()
        hora_creative["hora"] = hora_creative["hora"].apply(lambda x: f"{int(x):02d}:00")

        for c_idx, col_name in enumerate(hora_creative.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
        fmt_header_row(ws, r, len(hora_creative.columns))
        r += 1
        for _, row_data in hora_creative.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c_idx, value=val)
            r += 1
        fmt_data_area(ws, r - len(hora_creative), r - 1, len(hora_creative.columns))

        if len(hora_creative) > 1:
            chart = LineChart()
            chart.title = "Perfil Horario - CO2, Temperatura, Umidade"
            chart.style = 10
            chart.height = 15
            chart.width = 25
            data_start = r - len(hora_creative)
            data_end = r - 1
            for col_idx in [3, 4, 5]:
                ref = Reference(ws, min_col=col_idx, min_row=data_start, max_col=col_idx, max_row=data_end)
                chart.add_data(ref, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end)
            chart.set_categories(cats)
            ws.add_chart(chart, "K2")

    r2 = r + 2
    if not pca.empty:
        r2 = add_subtitle(ws, "Media Horaria - Ocupacao (PCA)", r2) if r < r2 else r2
        r2 += 1
        hora_pca = pca.groupby("hora")[
            ["pessoas", "pct_ocupacao", "temperatura", "umidade"]
        ].mean().round(2).reset_index()
        hora_pca["hora"] = hora_pca["hora"].apply(lambda x: f"{int(x):02d}:00")

        for c_idx, col_name in enumerate(hora_pca.columns, 1):
            ws.cell(row=r2, column=c_idx, value=col_name.replace("_", " ").title())
        fmt_header_row(ws, r2, len(hora_pca.columns))
        r2 += 1
        for _, row_data in hora_pca.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                ws.cell(row=r2, column=c_idx, value=val)
            r2 += 1
        fmt_data_area(ws, r2 - len(hora_pca), r2 - 1, len(hora_pca.columns))

    auto_width(ws, 6)

    # ==================================================================
    # SHEET 9 - CONFORMIDADE (THRESHOLD CHECK)
    # ==================================================================
    ws = wb.create_sheet("Conformidade")

    r = add_title(ws, "VERIFICACAO DE CONFORMIDADE VS LIMIARES")
    r += 1

    if not creative.empty:
        creative_check = creative.copy()
        creative_check["co2_acima"] = creative_check["co2"] > 1000
        creative_check["voc_acima"] = creative_check["voc"] > 400
        creative_check["temp_acima"] = creative_check["temperatura"] > 26
        creative_check["umid_acima"] = creative_check["umidade"] > 70
        creative_check["ruido_acima"] = creative_check["ruido"] > 70

        creative_check["co2_atencao"] = (creative_check["co2"] > 800) & (creative_check["co2"] <= 1000)
        creative_check["temp_atencao"] = (creative_check["temperatura"] > 24) & (creative_check["temperatura"] <= 26)

        r = add_subtitle(ws, "Percentual de Violacao por Parametro (Todos os Setores)", r)
        r += 1
        violation_pcts = []
        for param in ["co2", "voc", "temp", "umid", "ruido"]:
            crit = creative_check[f"{param}_acima"].mean() * 100
            aten = creative_check.get(f"{param}_atencao", pd.Series([0])).mean() * 100
            violation_pcts.append({
                "Parametro": param.upper(),
                "% Critico": round(crit, 1),
                "% Atencao": round(aten, 1),
                "% Normal": round(100 - crit - aten, 1),
            })
        df_viol = pd.DataFrame(violation_pcts)
        for c_idx, col_name in enumerate(df_viol.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name)
        fmt_header_row(ws, r, len(df_viol.columns))
        r += 1
        for _, row_data in df_viol.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER
                if c_idx == 2 and val > 10:
                    cell.fill = FILL_RED
                elif c_idx == 2 and val > 5:
                    cell.fill = FILL_YELLOW
            r += 1

        r += 2
        r = add_subtitle(ws, "Violacao por Setor - CO2 Critico (> 1000 ppm)", r)
        r += 1
        setor_viol = creative_check.groupby("nome_setor")["co2_acima"].mean().mul(100).round(1).reset_index()
        setor_viol.columns = ["Setor", "% Violacao CO2"]
        for c_idx, col_name in enumerate(setor_viol.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name)
        fmt_header_row(ws, r, 2)
        r += 1
        for _, row_data in setor_viol.iterrows():
            ws.cell(row=r, column=1, value=row_data["Setor"]).border = THIN_BORDER
            cell = ws.cell(row=r, column=2, value=row_data["% Violacao CO2"])
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if row_data["% Violacao CO2"] > 10:
                cell.fill = FILL_RED
            r += 1

    auto_width(ws, 5)

    # ==================================================================
    # SHEET 10 - OCUPACAO DETALHADA
    # ==================================================================
    ws = wb.create_sheet("Ocupacao")

    r = add_title(ws, "ANALISE DE OCUPACAO (PCA / HPD2)")
    r += 1

    if not pca.empty:
        pca_check = pca.copy()
        pca_check["superlotado"] = pca_check["pessoas"] > pca_check["capacidade_maxima"]
        pca_check["quase_lotado"] = (pca_check["pct_ocupacao"] >= 80) & (pca_check["pct_ocupacao"] < 100)

        r = add_subtitle(ws, "Ocupacao Media por Setor", r)
        r += 1
        occ = pca_check.groupby("nome_setor").agg(
            media_pessoas=("pessoas", "mean"),
            max_pessoas=("pessoas", "max"),
            media_pct=("pct_ocupacao", "mean"),
            max_pct=("pct_ocupacao", "max"),
            total_registros=("pessoas", "count"),
            vezes_superlotado=("superlotado", "sum"),
        ).round(1).reset_index()
        occ["pct_superlotacao"] = (occ["vezes_superlotado"] / occ["total_registros"] * 100).round(1)

        for c_idx, col_name in enumerate(occ.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
        fmt_header_row(ws, r, len(occ.columns))
        r += 1
        for _, row_data in occ.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER
                if "superlot" in occ.columns[c_idx - 1].lower() and isinstance(val, (int, float)) and val > 0:
                    cell.fill = FILL_RED
            r += 1

        r += 2
        r = add_subtitle(ws, "Distribuicao de Ocupacao por Hora", r)
        r += 1
        hora_occ = pca_check.groupby("hora").agg(
            media_pessoas=("pessoas", "mean"),
            max_pessoas=("pessoas", "max"),
            media_pct=("pct_ocupacao", "mean"),
        ).round(1).reset_index()
        hora_occ["hora"] = hora_occ["hora"].apply(lambda x: f"{int(x):02d}:00")

        for c_idx, col_name in enumerate(hora_occ.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name.replace("_", " ").title())
        fmt_header_row(ws, r, len(hora_occ.columns))
        r += 1
        for _, row_data in hora_occ.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c_idx, value=val).border = THIN_BORDER
            r += 1
        fmt_data_area(ws, r - len(hora_occ), r - 1, len(hora_occ.columns))

    auto_width(ws, 8)

    # ==================================================================
    # SALVAR
    # ==================================================================
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print("\nArquivo Excel gerado: {}".format(os.path.abspath(OUTPUT_PATH)))


if __name__ == "__main__":
    gerar_eda()
